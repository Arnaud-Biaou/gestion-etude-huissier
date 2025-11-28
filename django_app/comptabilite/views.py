from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q, F, Count
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.core.paginator import Paginator
from django.db import transaction
from decimal import Decimal
import json
import csv
import io
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta

from .models import (
    ExerciceComptable, CompteComptable, Journal, EcritureComptable,
    LigneEcriture, TypeOperation, ParametrageFiscal, DeclarationTVA,
    RapportComptable, ConfigurationComptable, Lettrage
)

# Imports conditionnels pour exports
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


@login_required
def dashboard_comptabilite(request):
    """Tableau de bord comptable"""
    config = ConfigurationComptable.get_instance()
    exercice = ExerciceComptable.get_exercice_courant()

    context = {
        'page_title': 'Comptabilité',
        'config': config,
        'exercice': exercice,
    }

    # Si pas configuré, rediriger vers l'assistant
    if not config.est_configure:
        return render(request, 'comptabilite/assistant_configuration.html', context)

    if exercice:
        # Calcul du chiffre d'affaires (classe 7)
        ca_query = LigneEcriture.objects.filter(
            compte__classe='7',
            ecriture__exercice=exercice,
            ecriture__statut='valide'
        )
        ca_annee = ca_query.aggregate(total=Coalesce(Sum('credit'), 0) - Coalesce(Sum('debit'), 0))['total']

        # CA du mois en cours
        today = timezone.now().date()
        debut_mois = today.replace(day=1)
        ca_mois = ca_query.filter(
            ecriture__date__gte=debut_mois
        ).aggregate(total=Coalesce(Sum('credit'), 0) - Coalesce(Sum('debit'), 0))['total']

        # Calcul des charges (classe 6)
        charges_query = LigneEcriture.objects.filter(
            compte__classe='6',
            ecriture__exercice=exercice,
            ecriture__statut='valide'
        )
        charges_annee = charges_query.aggregate(total=Coalesce(Sum('debit'), 0) - Coalesce(Sum('credit'), 0))['total']

        charges_mois = charges_query.filter(
            ecriture__date__gte=debut_mois
        ).aggregate(total=Coalesce(Sum('debit'), 0) - Coalesce(Sum('credit'), 0))['total']

        # Résultat
        resultat = ca_annee - charges_annee

        # Créances clients (compte 411)
        creances = LigneEcriture.objects.filter(
            compte__numero__startswith='411',
            ecriture__exercice=exercice,
            ecriture__statut='valide'
        ).aggregate(total=Coalesce(Sum('debit'), 0) - Coalesce(Sum('credit'), 0))['total']

        # Dettes fournisseurs (compte 401)
        dettes = LigneEcriture.objects.filter(
            compte__numero__startswith='401',
            ecriture__exercice=exercice,
            ecriture__statut='valide'
        ).aggregate(total=Coalesce(Sum('credit'), 0) - Coalesce(Sum('debit'), 0))['total']

        # Trésorerie (classe 5)
        tresorerie = LigneEcriture.objects.filter(
            compte__classe='5',
            ecriture__exercice=exercice,
            ecriture__statut='valide'
        ).aggregate(total=Coalesce(Sum('debit'), 0) - Coalesce(Sum('credit'), 0))['total']

        # Statistiques écritures
        nb_ecritures_mois = EcritureComptable.objects.filter(
            exercice=exercice,
            date__gte=debut_mois
        ).count()

        nb_brouillons = EcritureComptable.objects.filter(
            exercice=exercice,
            statut='brouillon'
        ).count()

        # Dernières écritures
        dernieres_ecritures = EcritureComptable.objects.filter(
            exercice=exercice
        ).order_by('-date', '-id')[:10]

        # Données pour le graphique mensuel
        mois_labels = []
        ca_mensuel = []
        charges_mensuel = []

        for i in range(6, 0, -1):
            mois_date = today - timedelta(days=30 * i)
            mois_debut = mois_date.replace(day=1)
            if mois_date.month == 12:
                mois_fin = mois_date.replace(year=mois_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                mois_fin = mois_date.replace(month=mois_date.month + 1, day=1) - timedelta(days=1)

            mois_labels.append(mois_date.strftime('%b'))

            ca_m = LigneEcriture.objects.filter(
                compte__classe='7',
                ecriture__exercice=exercice,
                ecriture__statut='valide',
                ecriture__date__gte=mois_debut,
                ecriture__date__lte=mois_fin
            ).aggregate(total=Coalesce(Sum('credit'), 0))['total']
            ca_mensuel.append(int(ca_m))

            ch_m = LigneEcriture.objects.filter(
                compte__classe='6',
                ecriture__exercice=exercice,
                ecriture__statut='valide',
                ecriture__date__gte=mois_debut,
                ecriture__date__lte=mois_fin
            ).aggregate(total=Coalesce(Sum('debit'), 0))['total']
            charges_mensuel.append(int(ch_m))

        context.update({
            'ca_annee': ca_annee,
            'ca_mois': ca_mois,
            'charges_annee': charges_annee,
            'charges_mois': charges_mois,
            'resultat': resultat,
            'creances': creances,
            'dettes': dettes,
            'tresorerie': tresorerie,
            'nb_ecritures_mois': nb_ecritures_mois,
            'nb_brouillons': nb_brouillons,
            'dernieres_ecritures': dernieres_ecritures,
            'mois_labels': json.dumps(mois_labels),
            'ca_mensuel': json.dumps(ca_mensuel),
            'charges_mensuel': json.dumps(charges_mensuel),
        })

    return render(request, 'comptabilite/dashboard.html', context)


@login_required
def saisie_ecriture(request):
    """Page de saisie des écritures avec les 3 modes"""
    config = ConfigurationComptable.get_instance()
    exercice = ExerciceComptable.get_exercice_courant()

    if not exercice:
        messages.error(request, "Aucun exercice comptable ouvert. Veuillez créer un exercice.")
        return redirect('comptabilite:dashboard')

    # Récupérer les types d'opérations pour le mode facile
    types_operations = TypeOperation.objects.filter(actif=True).order_by('ordre_affichage')

    # Récupérer les comptes et journaux pour les modes guidé et expert
    comptes = CompteComptable.objects.filter(actif=True).order_by('numero')
    journaux = Journal.objects.filter(actif=True)

    # Grouper les comptes par classe pour le mode guidé
    comptes_par_classe = {}
    for compte in comptes:
        classe_label = dict(CompteComptable.CLASSE_CHOICES).get(compte.classe, compte.classe)
        if classe_label not in comptes_par_classe:
            comptes_par_classe[classe_label] = []
        comptes_par_classe[classe_label].append(compte)

    context = {
        'page_title': 'Nouvelle écriture',
        'config': config,
        'exercice': exercice,
        'types_operations': types_operations,
        'comptes': comptes,
        'journaux': journaux,
        'comptes_par_classe': comptes_par_classe,
        'mode_defaut': config.mode_saisie_defaut,
        'today': timezone.now().date(),
    }

    return render(request, 'comptabilite/saisie_ecriture.html', context)


@require_POST
def api_creer_ecriture_facile(request):
    """API pour créer une écriture en mode facile"""
    try:
        data = json.loads(request.body)
        type_operation_code = data.get('type_operation')
        montant = Decimal(str(data.get('montant', 0)))
        date_str = data.get('date')
        description = data.get('description', '')

        if montant <= 0:
            return JsonResponse({'success': False, 'error': 'Le montant doit être positif'})

        type_operation = get_object_or_404(TypeOperation, code=type_operation_code)
        exercice = ExerciceComptable.get_exercice_courant()

        if not exercice:
            return JsonResponse({'success': False, 'error': 'Aucun exercice ouvert'})

        date_ecriture = datetime.strptime(date_str, '%Y-%m-%d').date()

        # Créer l'écriture
        numero = EcritureComptable.generer_numero(type_operation.journal, date_ecriture)
        ecriture = EcritureComptable.objects.create(
            numero=numero,
            date=date_ecriture,
            journal=type_operation.journal,
            exercice=exercice,
            libelle=f"{type_operation.libelle} - {description}",
            statut='brouillon',
            origine='manuelle'
        )

        # Créer les lignes
        LigneEcriture.objects.create(
            ecriture=ecriture,
            compte=type_operation.compte_debit,
            libelle=description or type_operation.libelle,
            debit=montant,
            credit=0
        )

        LigneEcriture.objects.create(
            ecriture=ecriture,
            compte=type_operation.compte_credit,
            libelle=description or type_operation.libelle,
            debit=0,
            credit=montant
        )

        return JsonResponse({
            'success': True,
            'ecriture_id': ecriture.id,
            'numero': ecriture.numero,
            'message': f'Écriture {numero} créée avec succès'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_POST
def api_creer_ecriture_guidee(request):
    """API pour créer une écriture en mode guidé"""
    try:
        data = json.loads(request.body)
        journal_id = data.get('journal')
        compte_debit_id = data.get('compte_debit')
        compte_credit_id = data.get('compte_credit')
        montant = Decimal(str(data.get('montant', 0)))
        date_str = data.get('date')
        libelle = data.get('libelle', '')

        if montant <= 0:
            return JsonResponse({'success': False, 'error': 'Le montant doit être positif'})

        journal = get_object_or_404(Journal, id=journal_id)
        compte_debit = get_object_or_404(CompteComptable, id=compte_debit_id)
        compte_credit = get_object_or_404(CompteComptable, id=compte_credit_id)
        exercice = ExerciceComptable.get_exercice_courant()

        if not exercice:
            return JsonResponse({'success': False, 'error': 'Aucun exercice ouvert'})

        date_ecriture = datetime.strptime(date_str, '%Y-%m-%d').date()

        # Créer l'écriture
        numero = EcritureComptable.generer_numero(journal, date_ecriture)
        ecriture = EcritureComptable.objects.create(
            numero=numero,
            date=date_ecriture,
            journal=journal,
            exercice=exercice,
            libelle=libelle,
            statut='brouillon',
            origine='manuelle'
        )

        # Créer les lignes
        LigneEcriture.objects.create(
            ecriture=ecriture,
            compte=compte_debit,
            libelle=libelle,
            debit=montant,
            credit=0
        )

        LigneEcriture.objects.create(
            ecriture=ecriture,
            compte=compte_credit,
            libelle=libelle,
            debit=0,
            credit=montant
        )

        return JsonResponse({
            'success': True,
            'ecriture_id': ecriture.id,
            'numero': ecriture.numero,
            'message': f'Écriture {numero} créée avec succès'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_POST
def api_creer_ecriture_expert(request):
    """API pour créer une écriture en mode expert (multi-lignes)"""
    try:
        data = json.loads(request.body)
        journal_id = data.get('journal')
        date_str = data.get('date')
        libelle = data.get('libelle', '')
        lignes = data.get('lignes', [])

        if not lignes or len(lignes) < 2:
            return JsonResponse({'success': False, 'error': 'Une écriture doit avoir au moins 2 lignes'})

        journal = get_object_or_404(Journal, id=journal_id)
        exercice = ExerciceComptable.get_exercice_courant()

        if not exercice:
            return JsonResponse({'success': False, 'error': 'Aucun exercice ouvert'})

        date_ecriture = datetime.strptime(date_str, '%Y-%m-%d').date()

        # Vérifier l'équilibre
        total_debit = sum(Decimal(str(l.get('debit', 0))) for l in lignes)
        total_credit = sum(Decimal(str(l.get('credit', 0))) for l in lignes)

        if total_debit != total_credit:
            return JsonResponse({
                'success': False,
                'error': f'L\'écriture n\'est pas équilibrée. Débit: {total_debit}, Crédit: {total_credit}'
            })

        # Créer l'écriture
        numero = EcritureComptable.generer_numero(journal, date_ecriture)
        ecriture = EcritureComptable.objects.create(
            numero=numero,
            date=date_ecriture,
            journal=journal,
            exercice=exercice,
            libelle=libelle,
            statut='brouillon',
            origine='manuelle'
        )

        # Créer les lignes
        for ligne_data in lignes:
            compte = get_object_or_404(CompteComptable, id=ligne_data.get('compte'))
            LigneEcriture.objects.create(
                ecriture=ecriture,
                compte=compte,
                libelle=ligne_data.get('libelle', libelle),
                debit=Decimal(str(ligne_data.get('debit', 0))),
                credit=Decimal(str(ligne_data.get('credit', 0)))
            )

        return JsonResponse({
            'success': True,
            'ecriture_id': ecriture.id,
            'numero': ecriture.numero,
            'message': f'Écriture {numero} créée avec succès'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_POST
def api_valider_ecriture(request):
    """Valider une écriture brouillon"""
    try:
        data = json.loads(request.body)
        ecriture_id = data.get('ecriture_id')

        ecriture = get_object_or_404(EcritureComptable, id=ecriture_id)

        if not ecriture.est_equilibree:
            return JsonResponse({
                'success': False,
                'error': f'L\'écriture n\'est pas équilibrée'
            })

        ecriture.valider()

        return JsonResponse({
            'success': True,
            'message': f'Écriture {ecriture.numero} validée'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def liste_ecritures(request):
    """Liste des écritures comptables"""
    exercice = ExerciceComptable.get_exercice_courant()

    ecritures = EcritureComptable.objects.all()

    # Filtres
    journal_filter = request.GET.get('journal')
    statut_filter = request.GET.get('statut')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    search = request.GET.get('search')

    if exercice:
        ecritures = ecritures.filter(exercice=exercice)

    if journal_filter:
        ecritures = ecritures.filter(journal_id=journal_filter)

    if statut_filter:
        ecritures = ecritures.filter(statut=statut_filter)

    if date_debut:
        ecritures = ecritures.filter(date__gte=date_debut)

    if date_fin:
        ecritures = ecritures.filter(date__lte=date_fin)

    if search:
        ecritures = ecritures.filter(
            Q(numero__icontains=search) |
            Q(libelle__icontains=search) |
            Q(reference__icontains=search)
        )

    ecritures = ecritures.order_by('-date', '-numero')

    # Pagination
    paginator = Paginator(ecritures, 25)
    page = request.GET.get('page')
    ecritures_page = paginator.get_page(page)

    context = {
        'page_title': 'Écritures comptables',
        'ecritures': ecritures_page,
        'journaux': Journal.objects.filter(actif=True),
        'exercice': exercice,
        'statuts': EcritureComptable.STATUT_CHOICES,
    }

    return render(request, 'comptabilite/liste_ecritures.html', context)


def detail_ecriture(request, ecriture_id):
    """Détail d'une écriture"""
    ecriture = get_object_or_404(EcritureComptable, id=ecriture_id)

    context = {
        'page_title': f'Écriture {ecriture.numero}',
        'ecriture': ecriture,
        'lignes': ecriture.lignes.all().select_related('compte'),
    }

    return render(request, 'comptabilite/detail_ecriture.html', context)


def journaux(request):
    """Liste des journaux comptables avec leurs écritures"""
    exercice = ExerciceComptable.get_exercice_courant()
    journal_id = request.GET.get('journal')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')

    journaux_list = Journal.objects.filter(actif=True)
    journal_selectionne = None
    ecritures = []

    if journal_id:
        journal_selectionne = get_object_or_404(Journal, id=journal_id)
        ecritures = EcritureComptable.objects.filter(
            journal=journal_selectionne,
            statut='valide'
        )

        if exercice:
            ecritures = ecritures.filter(exercice=exercice)

        if date_debut:
            ecritures = ecritures.filter(date__gte=date_debut)

        if date_fin:
            ecritures = ecritures.filter(date__lte=date_fin)

        ecritures = ecritures.order_by('date', 'numero').prefetch_related('lignes__compte')

    context = {
        'page_title': 'Journaux comptables',
        'journaux': journaux_list,
        'journal_selectionne': journal_selectionne,
        'ecritures': ecritures,
        'exercice': exercice,
    }

    return render(request, 'comptabilite/journaux.html', context)


def grand_livre(request):
    """Grand livre comptable"""
    exercice = ExerciceComptable.get_exercice_courant()
    compte_id = request.GET.get('compte')
    classe_filter = request.GET.get('classe')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')

    comptes = CompteComptable.objects.filter(actif=True).order_by('numero')

    if classe_filter:
        comptes = comptes.filter(classe=classe_filter)

    compte_selectionne = None
    mouvements = []
    solde_progressif = Decimal('0')

    if compte_id:
        compte_selectionne = get_object_or_404(CompteComptable, id=compte_id)
        lignes = LigneEcriture.objects.filter(
            compte=compte_selectionne,
            ecriture__statut='valide'
        ).select_related('ecriture', 'ecriture__journal')

        if exercice:
            lignes = lignes.filter(ecriture__exercice=exercice)

        if date_debut:
            lignes = lignes.filter(ecriture__date__gte=date_debut)

        if date_fin:
            lignes = lignes.filter(ecriture__date__lte=date_fin)

        lignes = lignes.order_by('ecriture__date', 'ecriture__numero')

        # Calculer le solde progressif
        for ligne in lignes:
            solde_progressif += ligne.debit - ligne.credit
            mouvements.append({
                'ligne': ligne,
                'solde': solde_progressif
            })

    context = {
        'page_title': 'Grand livre',
        'comptes': comptes,
        'compte_selectionne': compte_selectionne,
        'mouvements': mouvements,
        'exercice': exercice,
        'classes': CompteComptable.CLASSE_CHOICES,
    }

    return render(request, 'comptabilite/grand_livre.html', context)


def balance(request):
    """Balance générale des comptes"""
    exercice = ExerciceComptable.get_exercice_courant()
    classe_filter = request.GET.get('classe')
    date_fin = request.GET.get('date_fin', timezone.now().date())

    if isinstance(date_fin, str):
        date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()

    comptes = CompteComptable.objects.filter(actif=True).order_by('numero')

    if classe_filter:
        comptes = comptes.filter(classe=classe_filter)

    balance_data = []
    totaux = {'debit': 0, 'credit': 0, 'solde_debiteur': 0, 'solde_crediteur': 0}

    for compte in comptes:
        lignes = LigneEcriture.objects.filter(
            compte=compte,
            ecriture__statut='valide'
        )

        if exercice:
            lignes = lignes.filter(
                ecriture__exercice=exercice,
                ecriture__date__lte=date_fin
            )

        agg = lignes.aggregate(
            total_debit=Coalesce(Sum('debit'), Decimal('0')),
            total_credit=Coalesce(Sum('credit'), Decimal('0'))
        )

        total_debit = agg['total_debit']
        total_credit = agg['total_credit']
        solde = total_debit - total_credit

        if total_debit > 0 or total_credit > 0:
            balance_data.append({
                'compte': compte,
                'debit': total_debit,
                'credit': total_credit,
                'solde_debiteur': solde if solde > 0 else 0,
                'solde_crediteur': abs(solde) if solde < 0 else 0
            })

            totaux['debit'] += total_debit
            totaux['credit'] += total_credit
            totaux['solde_debiteur'] += solde if solde > 0 else 0
            totaux['solde_crediteur'] += abs(solde) if solde < 0 else 0

    context = {
        'page_title': 'Balance générale',
        'balance': balance_data,
        'totaux': totaux,
        'exercice': exercice,
        'classes': CompteComptable.CLASSE_CHOICES,
        'date_fin': date_fin,
    }

    return render(request, 'comptabilite/balance.html', context)


def etats_financiers(request):
    """États financiers OHADA (Bilan et Compte de résultat)"""
    exercice = ExerciceComptable.get_exercice_courant()
    type_etat = request.GET.get('type', 'bilan')

    context = {
        'page_title': 'États financiers',
        'exercice': exercice,
        'type_etat': type_etat,
    }

    if exercice:
        if type_etat == 'bilan':
            context['bilan'] = generer_bilan(exercice)
        elif type_etat == 'resultat':
            context['resultat'] = generer_compte_resultat(exercice)
        elif type_etat == 'flux':
            context['flux'] = generer_flux_tresorerie(exercice)

    return render(request, 'comptabilite/etats_financiers.html', context)


def generer_bilan(exercice):
    """Génère les données du bilan"""
    def get_solde_comptes(prefix_list):
        total = Decimal('0')
        for prefix in prefix_list:
            lignes = LigneEcriture.objects.filter(
                compte__numero__startswith=prefix,
                ecriture__exercice=exercice,
                ecriture__statut='valide'
            )
            agg = lignes.aggregate(
                debit=Coalesce(Sum('debit'), Decimal('0')),
                credit=Coalesce(Sum('credit'), Decimal('0'))
            )
            total += agg['debit'] - agg['credit']
        return total

    bilan = {
        'actif': {
            'immobilisations': abs(get_solde_comptes(['2'])),
            'stocks': abs(get_solde_comptes(['3'])),
            'creances_clients': abs(get_solde_comptes(['411'])),
            'autres_creances': abs(get_solde_comptes(['41']) - get_solde_comptes(['411'])),
            'banque': abs(get_solde_comptes(['52'])),
            'caisse': abs(get_solde_comptes(['57'])),
        },
        'passif': {
            'capital': abs(get_solde_comptes(['10'])),
            'reserves': abs(get_solde_comptes(['11'])),
            'resultat': 0,  # Calculé après
            'dettes_fournisseurs': abs(get_solde_comptes(['401'])),
            'dettes_fiscales': abs(get_solde_comptes(['44'])),
            'autres_dettes': abs(get_solde_comptes(['40']) - get_solde_comptes(['401'])),
        }
    }

    # Calcul du résultat
    produits = abs(get_solde_comptes(['7']))
    charges = abs(get_solde_comptes(['6']))
    bilan['passif']['resultat'] = produits - charges

    # Totaux
    bilan['total_actif'] = sum(bilan['actif'].values())
    bilan['total_passif'] = sum(bilan['passif'].values())

    return bilan


def generer_compte_resultat(exercice):
    """Génère les données du compte de résultat"""
    def get_solde_comptes(prefix_list):
        total = Decimal('0')
        for prefix in prefix_list:
            lignes = LigneEcriture.objects.filter(
                compte__numero__startswith=prefix,
                ecriture__exercice=exercice,
                ecriture__statut='valide'
            )
            agg = lignes.aggregate(
                debit=Coalesce(Sum('debit'), Decimal('0')),
                credit=Coalesce(Sum('credit'), Decimal('0'))
            )
            # Pour les produits, le solde est créditeur
            # Pour les charges, le solde est débiteur
            if prefix.startswith('7'):
                total += agg['credit'] - agg['debit']
            else:
                total += agg['debit'] - agg['credit']
        return total

    resultat = {
        'produits': {
            'honoraires': get_solde_comptes(['706']),
            'emoluments': get_solde_comptes(['707']),
            'autres_produits': get_solde_comptes(['70']) - get_solde_comptes(['706', '707']),
            'produits_financiers': get_solde_comptes(['76', '77']),
        },
        'charges': {
            'achats': get_solde_comptes(['60']),
            'services_exterieurs': get_solde_comptes(['61', '62']),
            'charges_personnel': get_solde_comptes(['64']),
            'impots_taxes': get_solde_comptes(['63']),
            'dotations_amortissements': get_solde_comptes(['68']),
            'charges_financieres': get_solde_comptes(['67']),
        }
    }

    resultat['total_produits'] = sum(resultat['produits'].values())
    resultat['total_charges'] = sum(resultat['charges'].values())
    resultat['resultat_net'] = resultat['total_produits'] - resultat['total_charges']

    return resultat


def generer_flux_tresorerie(exercice):
    """Génère un tableau simplifié des flux de trésorerie"""
    def get_variation(prefix_list):
        total = Decimal('0')
        for prefix in prefix_list:
            lignes = LigneEcriture.objects.filter(
                compte__numero__startswith=prefix,
                ecriture__exercice=exercice,
                ecriture__statut='valide'
            )
            agg = lignes.aggregate(
                debit=Coalesce(Sum('debit'), Decimal('0')),
                credit=Coalesce(Sum('credit'), Decimal('0'))
            )
            total += agg['debit'] - agg['credit']
        return total

    flux = {
        'exploitation': {
            'encaissements_clients': get_variation(['52', '57']),
            'decaissements_fournisseurs': 0,
        },
        'investissement': {
            'acquisitions': get_variation(['2']),
        },
        'financement': {
            'apports': get_variation(['10']),
        }
    }

    return flux


def plan_comptable(request):
    """Affichage et gestion du plan comptable"""
    classe_filter = request.GET.get('classe')
    search = request.GET.get('search')

    comptes = CompteComptable.objects.all().order_by('numero')

    if classe_filter:
        comptes = comptes.filter(classe=classe_filter)

    if search:
        comptes = comptes.filter(
            Q(numero__icontains=search) |
            Q(libelle__icontains=search)
        )

    # Grouper par classe
    comptes_par_classe = {}
    for compte in comptes:
        if compte.classe not in comptes_par_classe:
            comptes_par_classe[compte.classe] = []
        comptes_par_classe[compte.classe].append(compte)

    context = {
        'page_title': 'Plan comptable OHADA',
        'comptes_par_classe': comptes_par_classe,
        'classes': CompteComptable.CLASSE_CHOICES,
        'total_comptes': comptes.count(),
    }

    return render(request, 'comptabilite/plan_comptable.html', context)


def gestion_tva(request):
    """Gestion de la TVA et déclarations"""
    exercice = ExerciceComptable.get_exercice_courant()
    declarations = DeclarationTVA.objects.filter(exercice=exercice) if exercice else []

    # Calcul TVA courante
    tva_collectee = Decimal('0')
    tva_deductible = Decimal('0')

    if exercice:
        try:
            params = exercice.parametres_fiscaux
            if params.compte_tva_collectee:
                tva_collectee = LigneEcriture.objects.filter(
                    compte=params.compte_tva_collectee,
                    ecriture__exercice=exercice,
                    ecriture__statut='valide'
                ).aggregate(total=Coalesce(Sum('credit'), Decimal('0')))['total']

            if params.compte_tva_deductible:
                tva_deductible = LigneEcriture.objects.filter(
                    compte=params.compte_tva_deductible,
                    ecriture__exercice=exercice,
                    ecriture__statut='valide'
                ).aggregate(total=Coalesce(Sum('debit'), Decimal('0')))['total']
        except ParametrageFiscal.DoesNotExist:
            pass

    context = {
        'page_title': 'Gestion TVA',
        'exercice': exercice,
        'declarations': declarations,
        'tva_collectee': tva_collectee,
        'tva_deductible': tva_deductible,
        'tva_a_payer': tva_collectee - tva_deductible,
    }

    return render(request, 'comptabilite/gestion_tva.html', context)


def cloture_exercice(request):
    """Assistant de clôture d'exercice"""
    exercice = ExerciceComptable.get_exercice_courant()

    # Vérifications avant clôture
    verifications = []

    if exercice:
        # Écritures non validées
        nb_brouillons = EcritureComptable.objects.filter(
            exercice=exercice,
            statut='brouillon'
        ).count()
        verifications.append({
            'label': 'Écritures brouillon à valider',
            'valeur': nb_brouillons,
            'ok': nb_brouillons == 0,
            'message': f'{nb_brouillons} écriture(s) en brouillon' if nb_brouillons > 0 else 'Toutes les écritures sont validées'
        })

        # Balance équilibrée
        totaux = LigneEcriture.objects.filter(
            ecriture__exercice=exercice,
            ecriture__statut='valide'
        ).aggregate(
            total_debit=Coalesce(Sum('debit'), Decimal('0')),
            total_credit=Coalesce(Sum('credit'), Decimal('0'))
        )
        equilibre = totaux['total_debit'] == totaux['total_credit']
        verifications.append({
            'label': 'Équilibre de la balance',
            'valeur': f"D: {totaux['total_debit']} / C: {totaux['total_credit']}",
            'ok': equilibre,
            'message': 'Balance équilibrée' if equilibre else 'La balance n\'est pas équilibrée'
        })

    context = {
        'page_title': 'Clôture d\'exercice',
        'exercice': exercice,
        'verifications': verifications,
    }

    return render(request, 'comptabilite/cloture_exercice.html', context)


@require_POST
def api_configurer_comptabilite(request):
    """Configure le module comptabilité"""
    try:
        data = json.loads(request.body)

        # Créer l'exercice
        date_debut = datetime.strptime(data.get('date_debut'), '%Y-%m-%d').date()
        date_fin = datetime.strptime(data.get('date_fin'), '%Y-%m-%d').date()

        exercice = ExerciceComptable.objects.create(
            libelle=data.get('libelle', f'Exercice {date_debut.year}'),
            date_debut=date_debut,
            date_fin=date_fin,
            statut='ouvert',
            est_premier_exercice=True
        )

        # Créer les paramètres fiscaux
        ParametrageFiscal.objects.create(
            exercice=exercice,
            taux_tva=Decimal(str(data.get('taux_tva', 18)))
        )

        # Marquer comme configuré
        config = ConfigurationComptable.get_instance()
        config.est_configure = True
        config.date_configuration = timezone.now()
        config.save()

        return JsonResponse({
            'success': True,
            'message': 'Configuration terminée avec succès'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_GET
def api_comptes_par_classe(request, classe):
    """Retourne les comptes d'une classe spécifique"""
    comptes = CompteComptable.objects.filter(
        classe=classe,
        actif=True
    ).values('id', 'numero', 'libelle')

    return JsonResponse({'comptes': list(comptes)})


def aide_comptabilite(request):
    """Page d'aide et tutoriels"""
    context = {
        'page_title': 'Aide - Comptabilité',
    }
    return render(request, 'comptabilite/aide.html', context)


def rapports(request):
    """Page de génération des rapports"""
    exercice = ExerciceComptable.get_exercice_courant()

    # Récupérer les rapports sauvegardés
    rapports_sauvegardes = RapportComptable.objects.filter(
        exercice=exercice
    ).order_by('-date_generation')[:10] if exercice else []

    context = {
        'page_title': 'Rapports comptables',
        'exercice': exercice,
        'rapports_sauvegardes': rapports_sauvegardes,
    }

    return render(request, 'comptabilite/rapports.html', context)


# ============================================================================
# CORRECTION 7 : Balance âgée clients/fournisseurs
# ============================================================================
def balance_agee(request):
    """Balance âgée des comptes clients (411) et fournisseurs (401)"""
    exercice = ExerciceComptable.get_exercice_courant()
    type_compte = request.GET.get('type', 'clients')  # clients ou fournisseurs
    date_reference = request.GET.get('date', timezone.now().date())

    if isinstance(date_reference, str):
        date_reference = datetime.strptime(date_reference, '%Y-%m-%d').date()

    # Définir le préfixe de compte selon le type
    prefix = '411' if type_compte == 'clients' else '401'
    titre = 'Clients' if type_compte == 'clients' else 'Fournisseurs'

    # Récupérer les écritures non lettrées sur les comptes tiers
    lignes_query = LigneEcriture.objects.filter(
        compte__numero__startswith=prefix,
        ecriture__statut='valide'
    ).select_related('ecriture', 'compte')

    if exercice:
        lignes_query = lignes_query.filter(ecriture__exercice=exercice)

    # Exclure les lignes déjà lettrées (entièrement)
    lignes_query = lignes_query.exclude(
        lettrages__est_partiel=False
    )

    # Calculer l'âge de chaque ligne et regrouper
    balance_data = []
    totaux = {
        'non_echu': Decimal('0'),
        '0_30': Decimal('0'),
        '31_60': Decimal('0'),
        '61_90': Decimal('0'),
        'plus_90': Decimal('0'),
        'total': Decimal('0'),
    }

    # Grouper par tiers
    tiers_dict = {}
    for ligne in lignes_query:
        tiers = ligne.tiers or ligne.ecriture.libelle[:30]
        if tiers not in tiers_dict:
            tiers_dict[tiers] = {
                'tiers': tiers,
                'compte': ligne.compte,
                'non_echu': Decimal('0'),
                '0_30': Decimal('0'),
                '31_60': Decimal('0'),
                '61_90': Decimal('0'),
                'plus_90': Decimal('0'),
                'total': Decimal('0'),
            }

        # Calculer le solde (débit - crédit pour clients, crédit - débit pour fournisseurs)
        if type_compte == 'clients':
            solde = ligne.debit - ligne.credit
        else:
            solde = ligne.credit - ligne.debit

        # Calculer l'âge
        age = (date_reference - ligne.ecriture.date).days

        # Classer par tranche d'âge
        if age < 0:
            tiers_dict[tiers]['non_echu'] += solde
            totaux['non_echu'] += solde
        elif age <= 30:
            tiers_dict[tiers]['0_30'] += solde
            totaux['0_30'] += solde
        elif age <= 60:
            tiers_dict[tiers]['31_60'] += solde
            totaux['31_60'] += solde
        elif age <= 90:
            tiers_dict[tiers]['61_90'] += solde
            totaux['61_90'] += solde
        else:
            tiers_dict[tiers]['plus_90'] += solde
            totaux['plus_90'] += solde

        tiers_dict[tiers]['total'] += solde
        totaux['total'] += solde

    # Filtrer les tiers avec un solde non nul
    balance_data = [v for v in tiers_dict.values() if v['total'] != 0]
    balance_data.sort(key=lambda x: abs(x['total']), reverse=True)

    context = {
        'page_title': f'Balance âgée {titre}',
        'type_compte': type_compte,
        'titre': titre,
        'date_reference': date_reference,
        'balance_data': balance_data,
        'totaux': totaux,
        'exercice': exercice,
    }

    return render(request, 'comptabilite/balance_agee.html', context)


# ============================================================================
# CORRECTION 8 : Journal centralisateur
# ============================================================================
def journal_centralisateur(request):
    """Journal centralisateur - récapitulatif des totaux par journal"""
    exercice = ExerciceComptable.get_exercice_courant()
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')

    if not date_debut and exercice:
        date_debut = exercice.date_debut
    if not date_fin and exercice:
        date_fin = exercice.date_fin

    if isinstance(date_debut, str):
        date_debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
    if isinstance(date_fin, str):
        date_fin = datetime.strptime(date_fin, '%Y-%m-%d').date()

    journaux_data = []
    total_general_debit = Decimal('0')
    total_general_credit = Decimal('0')

    journaux = Journal.objects.filter(actif=True)

    for journal in journaux:
        ecritures = EcritureComptable.objects.filter(
            journal=journal,
            statut='valide'
        )

        if exercice:
            ecritures = ecritures.filter(exercice=exercice)
        if date_debut:
            ecritures = ecritures.filter(date__gte=date_debut)
        if date_fin:
            ecritures = ecritures.filter(date__lte=date_fin)

        # Calculer les totaux
        totaux = LigneEcriture.objects.filter(
            ecriture__in=ecritures
        ).aggregate(
            total_debit=Coalesce(Sum('debit'), Decimal('0')),
            total_credit=Coalesce(Sum('credit'), Decimal('0'))
        )

        nb_ecritures = ecritures.count()

        if nb_ecritures > 0:
            journaux_data.append({
                'journal': journal,
                'nb_ecritures': nb_ecritures,
                'total_debit': totaux['total_debit'],
                'total_credit': totaux['total_credit'],
            })

            total_general_debit += totaux['total_debit']
            total_general_credit += totaux['total_credit']

    context = {
        'page_title': 'Journal centralisateur',
        'journaux_data': journaux_data,
        'total_debit': total_general_debit,
        'total_credit': total_general_credit,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'exercice': exercice,
    }

    return render(request, 'comptabilite/journal_centralisateur.html', context)


# ============================================================================
# CORRECTION 9 : Intégration Facturation -> Comptabilité
# ============================================================================
def generer_ecriture_facture(facture, utilisateur=None):
    """
    Génère automatiquement l'écriture comptable pour une facture validée.

    Args:
        facture: Instance de gestion.Facture
        utilisateur: Utilisateur qui génère l'écriture

    Returns:
        EcritureComptable ou None si erreur
    """
    try:
        exercice = ExerciceComptable.get_exercice_courant()
        if not exercice:
            return None

        config = ConfigurationComptable.get_instance()

        # Déterminer le journal et les comptes selon le type de facture
        if facture.type_facture == 'honoraires':
            journal = Journal.objects.get(code='VE')
            compte_produit = CompteComptable.objects.filter(numero='7061').first()
        elif facture.type_facture == 'emoluments':
            journal = Journal.objects.get(code='VE')
            compte_produit = CompteComptable.objects.filter(numero='7062').first()
        else:
            journal = Journal.objects.get(code='VE')
            compte_produit = CompteComptable.objects.filter(numero='706').first()

        compte_client = config.compte_clients or CompteComptable.objects.filter(numero='411').first()

        if not compte_produit or not compte_client:
            return None

        # Créer l'écriture
        numero = EcritureComptable.generer_numero(journal, facture.date_facture)
        ecriture = EcritureComptable.objects.create(
            numero=numero,
            date=facture.date_facture,
            journal=journal,
            exercice=exercice,
            libelle=f"Facture {facture.numero} - {facture.client}",
            reference=facture.numero,
            statut='brouillon',
            origine='facture',
            facture=facture,
            cree_par=utilisateur
        )

        # Ligne client (débit)
        LigneEcriture.objects.create(
            ecriture=ecriture,
            compte=compte_client,
            libelle=f"Facture {facture.numero}",
            debit=facture.montant_ttc,
            credit=0,
            tiers=str(facture.client) if facture.client else ''
        )

        # Ligne produit (crédit HT)
        LigneEcriture.objects.create(
            ecriture=ecriture,
            compte=compte_produit,
            libelle=f"Facture {facture.numero}",
            debit=0,
            credit=facture.montant_ht
        )

        # Ligne TVA si applicable
        if facture.montant_tva and facture.montant_tva > 0:
            compte_tva = CompteComptable.objects.filter(numero='4431').first()
            if compte_tva:
                LigneEcriture.objects.create(
                    ecriture=ecriture,
                    compte=compte_tva,
                    libelle=f"TVA Facture {facture.numero}",
                    debit=0,
                    credit=facture.montant_tva
                )

        return ecriture

    except Exception as e:
        print(f"Erreur génération écriture facture: {e}")
        return None


# ============================================================================
# CORRECTION 10 : Intégration Trésorerie -> Comptabilité
# ============================================================================
def generer_ecriture_mouvement(mouvement, utilisateur=None):
    """
    Génère automatiquement l'écriture comptable pour un mouvement de trésorerie.

    Args:
        mouvement: Instance de tresorerie.MouvementTresorerie
        utilisateur: Utilisateur qui génère l'écriture

    Returns:
        EcritureComptable ou None si erreur
    """
    try:
        exercice = ExerciceComptable.get_exercice_courant()
        if not exercice:
            return None

        config = ConfigurationComptable.get_instance()

        # Déterminer le journal et le compte de trésorerie
        if mouvement.compte_bancaire.type_compte == 'caisse':
            journal = Journal.objects.get(code='CA')
            compte_tresorerie = config.compte_caisse or CompteComptable.objects.filter(numero='571').first()
        else:
            journal = Journal.objects.get(code='BQ')
            compte_tresorerie = config.compte_banque_principal or CompteComptable.objects.filter(numero='5211').first()

        if not compte_tresorerie:
            return None

        # Déterminer le compte de contrepartie selon le type de mouvement
        if mouvement.type_mouvement == 'encaissement':
            # Encaissement = débit trésorerie, crédit contrepartie
            if mouvement.facture:
                compte_contrepartie = CompteComptable.objects.filter(numero='411').first()
            else:
                compte_contrepartie = CompteComptable.objects.filter(numero='758').first()  # Produits divers
        else:
            # Décaissement = crédit trésorerie, débit contrepartie
            if mouvement.categorie == 'loyer':
                compte_contrepartie = CompteComptable.objects.filter(numero='6131').first()
            elif mouvement.categorie == 'salaire':
                compte_contrepartie = CompteComptable.objects.filter(numero='6411').first()
            elif mouvement.categorie == 'fournisseur':
                compte_contrepartie = CompteComptable.objects.filter(numero='401').first()
            else:
                compte_contrepartie = CompteComptable.objects.filter(numero='658').first()  # Charges diverses

        if not compte_contrepartie:
            return None

        # Créer l'écriture
        numero = EcritureComptable.generer_numero(journal, mouvement.date)
        ecriture = EcritureComptable.objects.create(
            numero=numero,
            date=mouvement.date,
            journal=journal,
            exercice=exercice,
            libelle=mouvement.libelle or f"Mouvement {mouvement.type_mouvement}",
            reference=mouvement.reference or '',
            statut='brouillon',
            origine='tresorerie',
            cree_par=utilisateur
        )

        # Créer les lignes selon le type
        if mouvement.type_mouvement == 'encaissement':
            LigneEcriture.objects.create(
                ecriture=ecriture,
                compte=compte_tresorerie,
                libelle=ecriture.libelle,
                debit=mouvement.montant,
                credit=0
            )
            LigneEcriture.objects.create(
                ecriture=ecriture,
                compte=compte_contrepartie,
                libelle=ecriture.libelle,
                debit=0,
                credit=mouvement.montant
            )
        else:
            LigneEcriture.objects.create(
                ecriture=ecriture,
                compte=compte_contrepartie,
                libelle=ecriture.libelle,
                debit=mouvement.montant,
                credit=0
            )
            LigneEcriture.objects.create(
                ecriture=ecriture,
                compte=compte_tresorerie,
                libelle=ecriture.libelle,
                debit=0,
                credit=mouvement.montant
            )

        return ecriture

    except Exception as e:
        print(f"Erreur génération écriture mouvement: {e}")
        return None


# ============================================================================
# CORRECTION 6 : Export PDF et Excel
# ============================================================================
def export_balance_pdf(request):
    """Export de la balance générale en PDF"""
    if not REPORTLAB_AVAILABLE:
        return JsonResponse({
            'success': False,
            'error': 'Le module reportlab n\'est pas installé. Installez-le avec: pip install reportlab'
        })

    exercice = ExerciceComptable.get_exercice_courant()
    if not exercice:
        return JsonResponse({'success': False, 'error': 'Aucun exercice ouvert'})

    # Récupérer les données de la balance
    comptes = CompteComptable.objects.filter(actif=True).order_by('numero')
    balance_data = []

    for compte in comptes:
        lignes = LigneEcriture.objects.filter(
            compte=compte,
            ecriture__statut='valide',
            ecriture__exercice=exercice
        )
        agg = lignes.aggregate(
            total_debit=Coalesce(Sum('debit'), Decimal('0')),
            total_credit=Coalesce(Sum('credit'), Decimal('0'))
        )

        if agg['total_debit'] > 0 or agg['total_credit'] > 0:
            solde = agg['total_debit'] - agg['total_credit']
            balance_data.append([
                compte.numero,
                compte.libelle[:40],
                f"{agg['total_debit']:,.0f}",
                f"{agg['total_credit']:,.0f}",
                f"{solde:,.0f}" if solde >= 0 else '',
                f"{abs(solde):,.0f}" if solde < 0 else '',
            ])

    # Créer le PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    styles = getSampleStyleSheet()

    # Titre
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        alignment=1,
        spaceAfter=30
    )
    elements.append(Paragraph(f"Balance Générale - {exercice.libelle}", title_style))
    elements.append(Paragraph(f"Au {timezone.now().strftime('%d/%m/%Y')}", styles['Normal']))
    elements.append(Spacer(1, 20))

    # Tableau
    header = ['Compte', 'Libellé', 'Débit', 'Crédit', 'Solde D', 'Solde C']
    data = [header] + balance_data

    table = Table(data, colWidths=[2*cm, 6*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="balance_{exercice.libelle}.pdf"'
    return response


def export_balance_excel(request):
    """Export de la balance générale en Excel"""
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({
            'success': False,
            'error': 'Le module openpyxl n\'est pas installé. Installez-le avec: pip install openpyxl'
        })

    exercice = ExerciceComptable.get_exercice_courant()
    if not exercice:
        return JsonResponse({'success': False, 'error': 'Aucun exercice ouvert'})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # En-têtes
    headers = ['Compte', 'Libellé', 'Débit', 'Crédit', 'Solde Débiteur', 'Solde Créditeur']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    # Données
    row = 2
    comptes = CompteComptable.objects.filter(actif=True).order_by('numero')
    totaux = {'debit': 0, 'credit': 0, 'solde_d': 0, 'solde_c': 0}

    for compte in comptes:
        lignes = LigneEcriture.objects.filter(
            compte=compte,
            ecriture__statut='valide',
            ecriture__exercice=exercice
        )
        agg = lignes.aggregate(
            total_debit=Coalesce(Sum('debit'), Decimal('0')),
            total_credit=Coalesce(Sum('credit'), Decimal('0'))
        )

        if agg['total_debit'] > 0 or agg['total_credit'] > 0:
            solde = agg['total_debit'] - agg['total_credit']
            ws.cell(row=row, column=1, value=compte.numero).border = border
            ws.cell(row=row, column=2, value=compte.libelle).border = border
            ws.cell(row=row, column=3, value=float(agg['total_debit'])).border = border
            ws.cell(row=row, column=4, value=float(agg['total_credit'])).border = border
            ws.cell(row=row, column=5, value=float(solde) if solde > 0 else 0).border = border
            ws.cell(row=row, column=6, value=float(abs(solde)) if solde < 0 else 0).border = border

            totaux['debit'] += agg['total_debit']
            totaux['credit'] += agg['total_credit']
            totaux['solde_d'] += solde if solde > 0 else 0
            totaux['solde_c'] += abs(solde) if solde < 0 else 0
            row += 1

    # Ligne totaux
    for col in range(1, 7):
        ws.cell(row=row, column=col).font = Font(bold=True)
        ws.cell(row=row, column=col).border = border
    ws.cell(row=row, column=1, value="TOTAUX")
    ws.cell(row=row, column=3, value=float(totaux['debit']))
    ws.cell(row=row, column=4, value=float(totaux['credit']))
    ws.cell(row=row, column=5, value=float(totaux['solde_d']))
    ws.cell(row=row, column=6, value=float(totaux['solde_c']))

    # Ajuster largeurs
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="balance_{exercice.libelle}.xlsx"'
    return response


def export_grand_livre_excel(request):
    """Export du grand livre en Excel"""
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({
            'success': False,
            'error': 'Le module openpyxl n\'est pas installé.'
        })

    exercice = ExerciceComptable.get_exercice_courant()
    compte_id = request.GET.get('compte')

    if not exercice:
        return JsonResponse({'success': False, 'error': 'Aucun exercice ouvert'})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grand Livre"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")

    # En-têtes
    headers = ['Date', 'Journal', 'N° Pièce', 'Libellé', 'Débit', 'Crédit', 'Solde']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    row = 2
    comptes_query = CompteComptable.objects.filter(actif=True).order_by('numero')
    if compte_id:
        comptes_query = comptes_query.filter(id=compte_id)

    for compte in comptes_query:
        lignes = LigneEcriture.objects.filter(
            compte=compte,
            ecriture__statut='valide',
            ecriture__exercice=exercice
        ).select_related('ecriture', 'ecriture__journal').order_by('ecriture__date')

        if lignes.exists():
            # Titre du compte
            ws.cell(row=row, column=1, value=f"{compte.numero} - {compte.libelle}")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1

            solde = Decimal('0')
            for ligne in lignes:
                solde += ligne.debit - ligne.credit
                ws.cell(row=row, column=1, value=ligne.ecriture.date.strftime('%d/%m/%Y'))
                ws.cell(row=row, column=2, value=ligne.ecriture.journal.code)
                ws.cell(row=row, column=3, value=ligne.ecriture.numero)
                ws.cell(row=row, column=4, value=ligne.libelle[:50])
                ws.cell(row=row, column=5, value=float(ligne.debit) if ligne.debit else '')
                ws.cell(row=row, column=6, value=float(ligne.credit) if ligne.credit else '')
                ws.cell(row=row, column=7, value=float(solde))
                row += 1
            row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="grand_livre_{exercice.libelle}.xlsx"'
    return response


def export_csv(request):
    """Export des données en CSV"""
    exercice = ExerciceComptable.get_exercice_courant()
    type_export = request.GET.get('type', 'ecritures')

    if not exercice:
        return JsonResponse({'success': False, 'error': 'Aucun exercice ouvert'})

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{type_export}_{exercice.libelle}.csv"'
    response.write('\ufeff')  # BOM UTF-8

    writer = csv.writer(response, delimiter=';')

    if type_export == 'ecritures':
        writer.writerow(['Date', 'Journal', 'N° Pièce', 'Libellé', 'Compte', 'Débit', 'Crédit'])

        lignes = LigneEcriture.objects.filter(
            ecriture__exercice=exercice,
            ecriture__statut='valide'
        ).select_related('ecriture', 'ecriture__journal', 'compte').order_by('ecriture__date')

        for ligne in lignes:
            writer.writerow([
                ligne.ecriture.date.strftime('%d/%m/%Y'),
                ligne.ecriture.journal.code,
                ligne.ecriture.numero,
                ligne.libelle,
                ligne.compte.numero,
                str(ligne.debit).replace('.', ',') if ligne.debit else '',
                str(ligne.credit).replace('.', ',') if ligne.credit else '',
            ])

    elif type_export == 'balance':
        writer.writerow(['Compte', 'Libellé', 'Débit', 'Crédit', 'Solde Débiteur', 'Solde Créditeur'])

        comptes = CompteComptable.objects.filter(actif=True).order_by('numero')
        for compte in comptes:
            lignes = LigneEcriture.objects.filter(
                compte=compte,
                ecriture__statut='valide',
                ecriture__exercice=exercice
            )
            agg = lignes.aggregate(
                total_debit=Coalesce(Sum('debit'), Decimal('0')),
                total_credit=Coalesce(Sum('credit'), Decimal('0'))
            )
            if agg['total_debit'] > 0 or agg['total_credit'] > 0:
                solde = agg['total_debit'] - agg['total_credit']
                writer.writerow([
                    compte.numero,
                    compte.libelle,
                    str(agg['total_debit']).replace('.', ','),
                    str(agg['total_credit']).replace('.', ','),
                    str(solde).replace('.', ',') if solde > 0 else '',
                    str(abs(solde)).replace('.', ',') if solde < 0 else '',
                ])

    return response


# ============================================================================
# CORRECTION 3 : Clôture d'exercice complète
# ============================================================================
@require_POST
def api_pre_cloture(request):
    """Lance la pré-clôture : génère les écritures d'inventaire"""
    try:
        exercice = ExerciceComptable.get_exercice_courant()
        if not exercice:
            return JsonResponse({'success': False, 'error': 'Aucun exercice ouvert'})

        # Vérifier qu'il n'y a pas de brouillons
        nb_brouillons = EcritureComptable.objects.filter(
            exercice=exercice,
            statut='brouillon'
        ).count()

        if nb_brouillons > 0:
            return JsonResponse({
                'success': False,
                'error': f'Il reste {nb_brouillons} écriture(s) en brouillon. Validez-les d\'abord.'
            })

        with transaction.atomic():
            journal_od = Journal.objects.get(code='OD')
            date_cloture = exercice.date_fin
            ecritures_generees = []

            # 1. Générer les écritures d'amortissement
            # Récupérer les immobilisations (classe 2)
            immobilisations = CompteComptable.objects.filter(
                numero__startswith='2',
                actif=True
            ).exclude(numero__startswith='28')  # Exclure les comptes d'amortissement

            compte_dotation = CompteComptable.objects.filter(numero='6811').first()

            for immo in immobilisations:
                solde = immo.get_solde(exercice=exercice)
                if solde > 0:
                    # Calculer l'amortissement (exemple: 20% linéaire)
                    taux = Decimal('0.20')
                    montant_amort = solde * taux

                    # Trouver le compte d'amortissement correspondant
                    numero_amort = '28' + immo.numero[1:]
                    compte_amort = CompteComptable.objects.filter(numero=numero_amort).first()

                    if compte_amort and compte_dotation and montant_amort > 0:
                        numero = EcritureComptable.generer_numero(journal_od, date_cloture)
                        ecriture = EcritureComptable.objects.create(
                            numero=numero,
                            date=date_cloture,
                            journal=journal_od,
                            exercice=exercice,
                            libelle=f"Dotation amortissement {immo.libelle}",
                            statut='brouillon',
                            origine='cloture'
                        )

                        LigneEcriture.objects.create(
                            ecriture=ecriture,
                            compte=compte_dotation,
                            libelle=f"Dotation amortissement {immo.numero}",
                            debit=montant_amort,
                            credit=0
                        )
                        LigneEcriture.objects.create(
                            ecriture=ecriture,
                            compte=compte_amort,
                            libelle=f"Amortissement {immo.numero}",
                            debit=0,
                            credit=montant_amort
                        )
                        ecritures_generees.append(ecriture.numero)

            return JsonResponse({
                'success': True,
                'message': f'Pré-clôture effectuée. {len(ecritures_generees)} écriture(s) d\'inventaire générée(s).',
                'ecritures': ecritures_generees
            })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@require_POST
def api_cloture_exercice(request):
    """Effectue la clôture définitive de l'exercice"""
    try:
        exercice = ExerciceComptable.get_exercice_courant()
        if not exercice:
            return JsonResponse({'success': False, 'error': 'Aucun exercice ouvert'})

        # Vérifications
        nb_brouillons = EcritureComptable.objects.filter(
            exercice=exercice,
            statut='brouillon'
        ).count()

        if nb_brouillons > 0:
            return JsonResponse({
                'success': False,
                'error': f'Il reste {nb_brouillons} écriture(s) en brouillon.'
            })

        with transaction.atomic():
            journal_cloture = Journal.objects.get_or_create(
                code='CL',
                defaults={
                    'libelle': 'Journal de clôture',
                    'type_journal': 'CL',
                    'actif': True
                }
            )[0]

            journal_an = Journal.objects.get_or_create(
                code='AN',
                defaults={
                    'libelle': 'À nouveau',
                    'type_journal': 'AN',
                    'actif': True
                }
            )[0]

            date_cloture = exercice.date_fin

            # 1. Calculer le résultat (produits - charges)
            produits = LigneEcriture.objects.filter(
                compte__classe='7',
                ecriture__exercice=exercice,
                ecriture__statut='valide'
            ).aggregate(
                total=Coalesce(Sum('credit'), Decimal('0')) - Coalesce(Sum('debit'), Decimal('0'))
            )['total']

            charges = LigneEcriture.objects.filter(
                compte__classe='6',
                ecriture__exercice=exercice,
                ecriture__statut='valide'
            ).aggregate(
                total=Coalesce(Sum('debit'), Decimal('0')) - Coalesce(Sum('credit'), Decimal('0'))
            )['total']

            resultat = produits - charges

            # 2. Écriture de détermination du résultat
            compte_resultat = CompteComptable.objects.filter(numero='130').first()
            if not compte_resultat:
                compte_resultat = CompteComptable.objects.filter(numero='13').first()

            if compte_resultat:
                numero = EcritureComptable.generer_numero(journal_cloture, date_cloture)
                ecriture_resultat = EcritureComptable.objects.create(
                    numero=numero,
                    date=date_cloture,
                    journal=journal_cloture,
                    exercice=exercice,
                    libelle="Détermination du résultat de l'exercice",
                    statut='valide',
                    origine='cloture'
                )

                # Solder les comptes de classe 6 et 7
                for classe in ['6', '7']:
                    comptes = CompteComptable.objects.filter(classe=classe, actif=True)
                    for compte in comptes:
                        solde = compte.get_solde(exercice=exercice)
                        if solde != 0:
                            if classe == '6':
                                LigneEcriture.objects.create(
                                    ecriture=ecriture_resultat,
                                    compte=compte,
                                    libelle=f"Solde {compte.numero}",
                                    debit=0 if solde > 0 else abs(solde),
                                    credit=solde if solde > 0 else 0
                                )
                            else:
                                LigneEcriture.objects.create(
                                    ecriture=ecriture_resultat,
                                    compte=compte,
                                    libelle=f"Solde {compte.numero}",
                                    debit=abs(solde) if solde < 0 else solde,
                                    credit=0 if solde < 0 else 0
                                )

                # Écriture au résultat
                if resultat >= 0:
                    LigneEcriture.objects.create(
                        ecriture=ecriture_resultat,
                        compte=compte_resultat,
                        libelle="Résultat de l'exercice (bénéfice)",
                        debit=0,
                        credit=resultat
                    )
                else:
                    LigneEcriture.objects.create(
                        ecriture=ecriture_resultat,
                        compte=compte_resultat,
                        libelle="Résultat de l'exercice (perte)",
                        debit=abs(resultat),
                        credit=0
                    )

            # 3. Clôturer l'exercice
            exercice.statut = 'cloture'
            exercice.save()

            # 4. Créer le nouvel exercice
            nouvel_exercice = ExerciceComptable.objects.create(
                libelle=f"Exercice {exercice.date_fin.year + 1}",
                date_debut=exercice.date_fin + timedelta(days=1),
                date_fin=date(exercice.date_fin.year + 1, exercice.date_fin.month, exercice.date_fin.day),
                statut='ouvert',
                est_premier_exercice=False
            )

            # 5. Générer les écritures de report à nouveau (classes 1-5)
            numero = EcritureComptable.generer_numero(journal_an, nouvel_exercice.date_debut)
            ecriture_an = EcritureComptable.objects.create(
                numero=numero,
                date=nouvel_exercice.date_debut,
                journal=journal_an,
                exercice=nouvel_exercice,
                libelle="Report à nouveau",
                statut='valide',
                origine='cloture'
            )

            for classe in ['1', '2', '3', '4', '5']:
                comptes = CompteComptable.objects.filter(classe=classe, actif=True)
                for compte in comptes:
                    solde = compte.get_solde(exercice=exercice)
                    if solde != 0:
                        LigneEcriture.objects.create(
                            ecriture=ecriture_an,
                            compte=compte,
                            libelle=f"À nouveau {compte.numero}",
                            debit=solde if solde > 0 else 0,
                            credit=abs(solde) if solde < 0 else 0
                        )

            return JsonResponse({
                'success': True,
                'message': f'Exercice clôturé. Nouvel exercice {nouvel_exercice.libelle} créé.',
                'nouvel_exercice_id': nouvel_exercice.id
            })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ============================================================================
# CORRECTION 14 : Flux de trésorerie détaillé
# ============================================================================
def generer_flux_tresorerie_detaille(exercice):
    """Génère un tableau détaillé des flux de trésorerie selon SYSCOHADA"""
    def get_variation(prefix_list, sens='debit'):
        total = Decimal('0')
        for prefix in prefix_list:
            lignes = LigneEcriture.objects.filter(
                compte__numero__startswith=prefix,
                ecriture__exercice=exercice,
                ecriture__statut='valide'
            )
            agg = lignes.aggregate(
                debit=Coalesce(Sum('debit'), Decimal('0')),
                credit=Coalesce(Sum('credit'), Decimal('0'))
            )
            if sens == 'debit':
                total += agg['debit'] - agg['credit']
            else:
                total += agg['credit'] - agg['debit']
        return total

    # Flux d'exploitation
    resultat_net = get_variation(['7'], 'credit') - get_variation(['6'], 'debit')
    amortissements = get_variation(['68'], 'debit')
    variation_stocks = get_variation(['3'], 'debit')
    variation_creances = get_variation(['41'], 'debit')
    variation_dettes_exploit = get_variation(['40'], 'credit')

    flux_exploitation = resultat_net + amortissements - variation_stocks - variation_creances + variation_dettes_exploit

    # Flux d'investissement
    acquisitions_immo = get_variation(['2'], 'debit')
    cessions_immo = get_variation(['82'], 'credit')
    flux_investissement = cessions_immo - acquisitions_immo

    # Flux de financement
    augmentation_capital = get_variation(['10'], 'credit')
    emprunts = get_variation(['16'], 'credit')
    remboursements = get_variation(['16'], 'debit')
    flux_financement = augmentation_capital + emprunts - remboursements

    # Variation de trésorerie
    variation_tresorerie = flux_exploitation + flux_investissement + flux_financement

    # Trésorerie début/fin
    tresorerie_debut = Decimal('0')  # À calculer si exercice N-1 existe
    tresorerie_fin = tresorerie_debut + variation_tresorerie

    return {
        'exploitation': {
            'resultat_net': resultat_net,
            'amortissements': amortissements,
            'variation_stocks': variation_stocks,
            'variation_creances': variation_creances,
            'variation_dettes': variation_dettes_exploit,
            'total': flux_exploitation,
        },
        'investissement': {
            'acquisitions': acquisitions_immo,
            'cessions': cessions_immo,
            'total': flux_investissement,
        },
        'financement': {
            'augmentation_capital': augmentation_capital,
            'emprunts': emprunts,
            'remboursements': remboursements,
            'total': flux_financement,
        },
        'variation_tresorerie': variation_tresorerie,
        'tresorerie_debut': tresorerie_debut,
        'tresorerie_fin': tresorerie_fin,
    }


# ============================================================================
# CORRECTION 16 : Consultation rapports sauvegardés
# ============================================================================
def detail_rapport(request, rapport_id):
    """Affiche le détail d'un rapport comptable sauvegardé"""
    rapport = get_object_or_404(RapportComptable, id=rapport_id)

    context = {
        'page_title': f'{rapport.get_type_rapport_display()} - {rapport.periode_fin}',
        'rapport': rapport,
    }

    return render(request, 'comptabilite/detail_rapport.html', context)


# ============================================================================
# CORRECTION 13 : API Déclarations TVA
# ============================================================================
@require_POST
def api_creer_declaration_tva(request):
    """Crée une nouvelle déclaration de TVA"""
    try:
        data = json.loads(request.body)
        exercice = ExerciceComptable.get_exercice_courant()

        if not exercice:
            return JsonResponse({'success': False, 'error': 'Aucun exercice ouvert'})

        periode_debut = datetime.strptime(data.get('periode_debut'), '%Y-%m-%d').date()
        periode_fin = datetime.strptime(data.get('periode_fin'), '%Y-%m-%d').date()

        declaration = DeclarationTVA.objects.create(
            exercice=exercice,
            periode_debut=periode_debut,
            periode_fin=periode_fin,
            statut='brouillon'
        )

        # Calculer les montants
        declaration.calculer()

        return JsonResponse({
            'success': True,
            'declaration_id': declaration.id,
            'message': 'Déclaration TVA créée'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def detail_declaration_tva(request, declaration_id):
    """Affiche le détail d'une déclaration TVA"""
    declaration = get_object_or_404(DeclarationTVA, id=declaration_id)

    context = {
        'page_title': f'Déclaration TVA {declaration.periode_debut.strftime("%m/%Y")}',
        'declaration': declaration,
    }

    return render(request, 'comptabilite/detail_declaration_tva.html', context)


# ============================================================================
# Vue Lettrage
# ============================================================================
def lettrage(request):
    """Page de lettrage des comptes tiers"""
    exercice = ExerciceComptable.get_exercice_courant()
    compte_id = request.GET.get('compte')
    type_compte = request.GET.get('type', 'clients')

    prefix = '411' if type_compte == 'clients' else '401'
    comptes = CompteComptable.objects.filter(
        numero__startswith=prefix,
        actif=True
    ).order_by('numero')

    lignes_non_lettrees = []
    compte_selectionne = None

    if compte_id:
        compte_selectionne = get_object_or_404(CompteComptable, id=compte_id)
        lignes_non_lettrees = LigneEcriture.objects.filter(
            compte=compte_selectionne,
            ecriture__statut='valide'
        ).exclude(
            lettrages__est_partiel=False
        ).select_related('ecriture').order_by('ecriture__date')

        if exercice:
            lignes_non_lettrees = lignes_non_lettrees.filter(ecriture__exercice=exercice)

    # Récupérer les lettrages existants
    lettrages_recents = Lettrage.objects.all().order_by('-date_lettrage')[:20]

    context = {
        'page_title': 'Lettrage des comptes',
        'type_compte': type_compte,
        'comptes': comptes,
        'compte_selectionne': compte_selectionne,
        'lignes_non_lettrees': lignes_non_lettrees,
        'lettrages_recents': lettrages_recents,
        'exercice': exercice,
    }

    return render(request, 'comptabilite/lettrage.html', context)


@require_POST
def api_creer_lettrage(request):
    """API pour créer un lettrage"""
    try:
        data = json.loads(request.body)
        compte_id = data.get('compte_id')
        ligne_ids = data.get('lignes', [])
        commentaire = data.get('commentaire', '')

        if not ligne_ids or len(ligne_ids) < 2:
            return JsonResponse({
                'success': False,
                'error': 'Sélectionnez au moins 2 lignes à lettrer'
            })

        compte = get_object_or_404(CompteComptable, id=compte_id)
        lignes = LigneEcriture.objects.filter(id__in=ligne_ids)

        # Vérifier que toutes les lignes appartiennent au même compte
        for ligne in lignes:
            if ligne.compte_id != compte.id:
                return JsonResponse({
                    'success': False,
                    'error': 'Toutes les lignes doivent appartenir au même compte'
                })

        utilisateur = request.user if hasattr(request, 'user') else None
        lettrage = Lettrage.creer_lettrage(compte, lignes, utilisateur, commentaire)

        return JsonResponse({
            'success': True,
            'lettrage_code': lettrage.code,
            'message': f'Lettrage {lettrage.code} créé'
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
